""" 
1 - Crie uma classe para representar um jogador de futebol, com os atributos nome, posição, data de nascimento, nacionalidade, altura e peso. Crie os métodos públicos necessários para sets e gets e também um método para imprimir todos os dados do jogador. Crie um método para calcular a idade do jogador e outro método para mostrar quanto tempo falta para o jogador se aposentar. Para isso, considere que os jogadores da posição de defesa se aposentam em média aos 40 anos, os jogadores de meio-campo aos 38 e os atacantes aos 35. 
"""

import datetime as dt
from tkinter import *
from tkinter.ttk import Combobox
import openpyxl as xl


class Jogador:
    def __init__(self):
        self.nome=''
        self.data_nasc=0
        self.nacionalidade=''
        self.nacoes = ['Brasil', 'Argentina', 'Uruguai', 'Paraguai', 'Chile', 'Bolívia', 'Colômbia']
        self.posicoes = ['defesa', 'meio-campo', 'ataque']
        self.altura=0
        self.peso=0
        self.posicao=''
        self.idade_aposentadoria={'defesa':40, 'meio-campo':38, 'ataque':35}
        self.is_aposentado = None

    def calcular_idade(self):
        return 2022 - self.data_nasc
    def calcular_aposentadoria(self):
        dif = self.idade_aposentadoria[self.posicao] - self.calcular_idade()
        if dif <=0:
            self.is_aposentado = True
        else:
            self.is_aposentado = False
        return dif

    def get_val_aposentadoria(self):
        return self.idade_aposentadoria[self.posicao]
    def get_is_aposentado(self):
        return self.is_aposentado 

    def get_attrs(self):
        return [
            self.get_nome(), 
            self.get_data_nasc(),
            self.get_peso(),
            self.get_altura(),
            self.get_nacionalidade(),
            self.get_posicao(),
            self.get_is_aposentado()
        ]

    def set_nome(self, nome):
        self.nome=nome
    def get_nome(self):
        return self.nome

    def set_data_nasc(self, data_nasc):
        self.data_nasc=data_nasc
    def get_data_nasc(self):
        return self.data_nasc

    def set_nacionalidade(self, nacionalidade):
        self.nacionalidade=nacionalidade
    def get_nacionalidade(self):
        return self.nacionalidade

    def set_altura(self, altura):
        self.altura=altura
    def get_altura(self):
        return self.altura

    def set_peso(self, peso):
        self.peso=peso
    def get_peso(self):
        return self.peso

    def set_posicao(self, posicao):
        self.posicao=posicao
    def get_posicao(self):
        return self.posicao

    

""" pedro = Jogador('Pedro', 2004, 'Brasileiro', 1.83, 83, 'meio-campo')
print(pedro.calcular_aposentadoria())
print(pedro.calcular_idade())
print(pedro.get_val_aposentadoria()) """

class Janela:
    
    def __init__(self, root: Tk):
        self.root = root
        self.nome_jogador = StringVar()
        self.dt_nasc_jogador = IntVar()
        self.altura_jogador = DoubleVar()
        self.peso_jogador = DoubleVar()
        self.nacao_jogador = StringVar()
        self.posicao_jogador = StringVar()
        self.jogadores = {}

        self.ety_nome = Entry(root, textvariable=self.nome_jogador)
        Label(root, text='Nome:').grid(column=0, row=0)
        self.ety_nome.grid(column=1, row=0)

        self.ety_data_nasc = Entry(root, textvariable=self.dt_nasc_jogador)
        Label(root, text='Ano de nascimento:').grid(column=0, row=1)
        self.ety_data_nasc.grid(column=1, row=1)

        self.ety_altura = Entry(root, textvariable=self.altura_jogador)
        Label(root, text='Altura:').grid(column=0, row=2)
        self.ety_altura.grid(column=1, row=2)

        self.ety_peso = Entry(root, textvariable=self.peso_jogador)
        Label(root, text='Peso:').grid(column=0, row=3)
        self.ety_peso.grid(column=1, row=3)

        self.cbb_posicao = Combobox(root, textvariable=self.posicao_jogador, values=Jogador().posicoes)
        Label(root, text='Posição:').grid(column=0, row=4)
        self.cbb_posicao.grid(column=1, row=4)

        self.cbb_nacao = Combobox(root, textvariable=self.nacao_jogador, values=Jogador().nacoes)
        Label(root, text='Nação:').grid(column=0, row=5)
        self.cbb_nacao.grid(column=1, row=5)

        self.bt_enviar = Button(root, text='Enviar Dados', command=self.enviar_dados)
        self.bt_enviar.grid(column=0, row=6, pady=20)

    def enviar_dados(self):

        jogador = Jogador()
        
        jogador.set_nome(self.nome_jogador.get())
        jogador.set_data_nasc(self.dt_nasc_jogador.get())
        jogador.set_altura(self.altura_jogador.get())
        jogador.set_peso(self.peso_jogador.get())
        jogador.set_nacionalidade(self.nacao_jogador.get())
        jogador.set_posicao(self.posicao_jogador.get())
        jogador.calcular_aposentadoria()

        book = xl.Workbook()
        ws = book['Sheet']
        ws.title = 'Jogadores'
        ws['A1'] = 'Nome'
        ws['B1'] = 'Data Nascimento'
        ws['C1'] = 'Altura'
        ws['D1'] = 'Peso'
        ws['E1'] = 'Nacionalidade'
        ws['F1'] = 'Posição'
        ws['G1'] = 'Idade Aposentadoria'
        ws['H1'] = 'Pode se aposentar'

        self.jogadores[jogador.get_nome()] = jogador
        print(self.jogadores.keys())

        for i in self.jogadores:
            ws.append(self.jogadores[i].get_attrs())

        

        book.save('Jogadores.xlsx')

    def set_settings(self, width, height, title):
        self.root.geometry(str(width)+'x'+str(height))
        self.root.title(title)

    def run(self):
        self.root.mainloop()


janela = Janela(Tk())
janela.set_settings(500,500,'teste')
janela.run()

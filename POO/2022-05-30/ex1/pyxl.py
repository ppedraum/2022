import openpyxl as xl

book = xl.Workbook() #cria uma 'área de trabalho"
ws = book['Sheet']
ws.title = 'Compras'
print(book.chartsheets)

ws['A1'] = 'Frutas'
ws['B1'] = 'Qtd.'
ws.append(['Banana', '45'])

book.save('Compras.xlsx')


from tkinter import *
import openpyxl as xl

book = xl.Workbook()
ws = book['Sheet']
ws.title = 'Lista de Compras'

root = Tk()
root.title('Lista de Compras')
root.geometry('500x500')

fr_user = Frame(root)
fr_user.grid(row=1, column=1)

fr_output = Frame(root)
fr_output.grid(row=1, column=2)

itens = []

def enviar():
    itens.append([item.get(), qtd.get()])

def salvar():
    ws['A1'] = 'Tipo'
    ws['B1'] = 'Quantidade'
    for i in itens:
        ws.append(i)
    book.save('Compras.xlsx')
    print(len(itens))
    print(itens)

def listar():
    lista_itens = ''
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        lista_itens+= f'Tipo: {row[0]} | Quantidade: {row[1]} \n'
    lbl_itens['text'] = lista_itens
     



item = StringVar()
Label(fr_user, text='Tipo de Item: ').grid(row=1, column=1)
Entry(fr_user, text='Insira', textvariable=item).grid(row=1, column=2)

qtd = StringVar()
Label(fr_user, text='Quantidade: ').grid(row=2, column=1)
Entry(fr_user, text='Insira', textvariable=qtd).grid(row=2, column=2)

Button(fr_user, text='Enviar', command=enviar).grid(row=3, column=1)
Button(fr_user, text='Salvar', command=salvar).grid(row=3, column=2)
Button(fr_user, text='Listar', command=listar).grid(row=3, column=3)

lbl_itens = Label(fr_output, text='')
lbl_itens.grid(row=1, column=1)


root.mainloop()
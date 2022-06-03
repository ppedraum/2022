import random as rd
import string
import openpyxl as xl



def ger_senhas():
    gerador = ''.join(rd.choice(string.ascii_uppercase + string.digits + string.punctuation) for i in range(rd.randint(1, 20)))
    return gerador

def run():
    book = xl.Workbook()
    ws = book['Sheet']
    ws.title = 'Senhas'

    ws['A1'] = 'Endereco'
    ws['B1'] = 'Senha'

    while True:
        endereco = input('Digite o endereço: ')
        cond = input('Senha automática (Y/N): ').upper()
        if cond == 'Y':
            senha = ger_senhas()
        elif cond == 'N':
            senha = input('Digite sua Senha: ')
        ws.append([endereco, senha])
        cond = input('Deseja continuar? (Y/N): ').upper()
        if cond == 'Y':
            continue
        elif cond == 'N':
            break
    book.save('Senhas.xlsx')

run()
